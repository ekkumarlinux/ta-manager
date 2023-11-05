import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import os
import argparse

def log(message):
    timestamp = datetime.now().strftime("%b %d, %Y, %I:%M:%S %p")
    print(f"{timestamp}\nInfo\n{message}")

def mark_and_delete_red_rows(filename, account_ids):
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        print(f"File not found: {filename}. Make sure the file exists in the specified path.")
        return

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows_to_delete = []

        for row in sheet.iter_rows():
            for i, cell in enumerate(row[:3]):  # Check only the first 3 columns
                cell_value = str(cell.value) if cell.value is not None else ""
                if any(account_id in cell_value for account_id in account_ids):
                    log(f"Account ID {cell_value} marked red on sheet: {sheet_name}")
                    for cell in row:
                        cell.fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                    rows_to_delete.append(row)
                    break

        for row in reversed(rows_to_delete):
            sheet.delete_rows(row[0].row)
            log(f"Account ID {cell_value} row removed from sheet: {sheet_name}")

    workbook.save(filename)
    log("Red rows have been marked and deleted in the Excel workbook.")

def main():
    parser = argparse.ArgumentParser(description="Mark and delete red rows in Excel files.")
    parser.add_argument("filename", help="Excel file name located in the 'reports/' directory")
    parser.add_argument("account_ids", help="Account IDs (comma-separated)")

    args = parser.parse_args()
    excel_file = os.path.join("reports", args.filename)
    account_ids = args.account_ids.split(',')

    mark_and_delete_red_rows(excel_file, account_ids)

if __name__ == "__main__":
    main()
